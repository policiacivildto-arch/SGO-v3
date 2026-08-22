import json
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError

from api.models import RosterPolicial

MATRICULA_KEYS = ('matricula', 'mat')
NOME_KEYS = ('nome', 'nome_completo', 'nome completo')
CARGO_KEYS = ('cargo_atual', 'cargo atual', 'cargo')


def _normalize_header(value):
    return str(value or '').strip().lower()


def _find_column(headers, candidate_keys):
    normalized = {_normalize_header(h): idx for idx, h in enumerate(headers)}
    for key in candidate_keys:
        if key in normalized:
            return normalized[key]
    return None


def _records_from_json(path):
    try:
        with open(path, encoding='utf-8') as f:
            data = json.load(f)
    except json.JSONDecodeError as exc:
        raise CommandError(f"JSON inválido em {path}: {exc}")

    for item in data:
        yield {
            'matricula': item.get('matricula', ''),
            'nome': item.get('nome', ''),
            'cargo': item.get('cargo_atual', item.get('cargo', '')),
        }


def _records_from_xlsx(path):
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise CommandError(
            "openpyxl não está instalado. Rode: pip install -r requirements.txt"
        )

    workbook = load_workbook(filename=path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)

    try:
        headers = next(rows)
    except StopIteration:
        raise CommandError(f"Planilha vazia: {path}")

    matricula_col = _find_column(headers, MATRICULA_KEYS)
    nome_col = _find_column(headers, NOME_KEYS)
    cargo_col = _find_column(headers, CARGO_KEYS)

    if matricula_col is None or nome_col is None:
        raise CommandError(
            "Não encontrei colunas de matrícula/nome na planilha. "
            f"Cabeçalhos lidos: {list(headers)}. "
            "Renomeie as colunas para 'matricula' e 'nome' (cargo é opcional) e tente de novo."
        )

    for row in rows:
        yield {
            'matricula': row[matricula_col] if matricula_col < len(row) else '',
            'nome': row[nome_col] if nome_col < len(row) else '',
            'cargo': row[cargo_col] if (cargo_col is not None and cargo_col < len(row)) else '',
        }


class Command(BaseCommand):
    help = (
        "Popula RosterPolicial (matricula/nome/cargo) a partir de um arquivo .json ou "
        ".xlsx com os dados do efetivo (colunas/chaves: matricula, nome, cargo). "
        "Usado para o autocompletar de matrícula nos formulários do frontend. "
        "O arquivo de origem NÃO é versionado no repositório (contém dados pessoais do "
        "efetivo) — informe o caminho local via --path."
    )

    def add_arguments(self, parser):
        parser.add_argument(
            '--path',
            required=True,
            help='Caminho do arquivo .json ou .xlsx de origem com os dados do efetivo.',
        )

    def handle(self, *args, **options):
        path = options['path']
        suffix = Path(path).suffix.lower()

        if suffix == '.json':
            record_source = _records_from_json(path)
        elif suffix in ('.xlsx', '.xlsm'):
            record_source = _records_from_xlsx(path)
        else:
            raise CommandError(f"Formato não suportado: {suffix}. Use .json ou .xlsx.")

        try:
            records = list(record_source)
        except FileNotFoundError:
            raise CommandError(f"Arquivo não encontrado: {path}")

        criados = 0
        atualizados = 0
        ignorados = 0

        for item in records:
            matricula = str(item.get('matricula') or '').strip()
            nome = str(item.get('nome') or '').strip()
            cargo = str(item.get('cargo') or '').strip()

            if not matricula or not nome:
                ignorados += 1
                continue

            _, created = RosterPolicial.objects.update_or_create(
                matricula=matricula,
                defaults={'nome': nome, 'cargo': cargo},
            )
            if created:
                criados += 1
            else:
                atualizados += 1

        self.stdout.write(self.style.SUCCESS(
            f"Roster atualizado: {criados} criados, {atualizados} atualizados, "
            f"{ignorados} ignorados (sem matrícula/nome)."
        ))
